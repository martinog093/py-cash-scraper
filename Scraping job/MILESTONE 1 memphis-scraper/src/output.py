import logging
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

COLUMNS = [
    "filing_date",
    "record_type",
    "primary_name",
    "secondary_name",
    "attorney_name",
    "docket_number",
    "verified_address",
    "unverified_address",
    "parcel_id",
    "debt_amount",
    "status",
    "discard_reason",
]


def write_output(records: list[dict], filepath: str) -> None:
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # Normalise records to output schema
    rows = []
    for r in records:
        rows.append({
            "filing_date": r.get("filing_date", ""),
            "record_type": r.get("record_type", ""),
            "primary_name": r.get("primary_name", ""),
            "secondary_name": r.get("secondary_name", ""),
            "attorney_name": r.get("attorney_name", ""),
            "docket_number": r.get("docket_number", ""),
            "verified_address": r.get("verified_address", ""),
            "unverified_address": r.get("unverified_address", ""),
            "parcel_id": r.get("parcel_id", ""),
            "debt_amount": r.get("debt_amount", ""),
            "status": r.get("status", ""),
            "discard_reason": r.get("discard_reason", ""),
        })

    new_df = pd.DataFrame(rows, columns=COLUMNS)

    existing_df = _read_existing_records(filepath)
    if not existing_df.empty:
        before = len(existing_df)
        df = _dedupe_by_docket(pd.concat([existing_df, new_df], ignore_index=True))
        logger.info(
            "Merged %d new record(s) with %d existing — %d total after de-duplicating by docket number",
            len(new_df), before, len(df),
        )
    else:
        df = new_df

    verified_df = df[df["status"] == "Verified"]
    review_df = df[df["status"].isin(["Needs Review", "Assessor Unavailable", "Docket Unavailable"])]
    discarded_df = df[df["status"] == "Discarded"]

    filepath = _resolve_writable_path(filepath)

    verified_df.to_excel(filepath, index=False, sheet_name="Verified Leads")

    wb = load_workbook(filepath)

    # Format the default sheet (Verified Leads, written by to_excel above)
    ws_verified = wb["Verified Leads"]
    _format_sheet(ws_verified, verified_df)

    # Add Needs Review sheet
    ws_review = wb.create_sheet("Needs Review")
    _write_dataframe_to_sheet(ws_review, review_df)
    _format_sheet(ws_review, review_df)

    # Add Discarded sheet
    ws_discarded = wb.create_sheet("Discarded")
    _write_dataframe_to_sheet(ws_discarded, discarded_df)
    _format_sheet(ws_discarded, discarded_df)

    wb.save(filepath)
    logger.info(
        "Output written to %s — %d verified, %d needs review, %d discarded",
        filepath,
        len(verified_df),
        len(review_df),
        len(discarded_df),
    )


def _read_existing_records(filepath: str) -> pd.DataFrame:
    """
    Load the previous run's results from filepath (all three sheets) so this
    run's records can be merged in rather than replacing the file outright.
    Returns an empty DataFrame if there's no previous file, or it can't be
    read (e.g. unexpected format) — in which case this run's results simply
    become the whole file, as before.
    """
    if not os.path.exists(filepath):
        return pd.DataFrame(columns=COLUMNS)

    try:
        sheets = pd.read_excel(filepath, sheet_name=["Verified Leads", "Needs Review", "Discarded"])
    except Exception as e:
        logger.warning("Could not read existing output %s for merging — starting fresh: %s", filepath, e)
        return pd.DataFrame(columns=COLUMNS)

    frames = [sheet.reindex(columns=COLUMNS).fillna("") for sheet in sheets.values()]
    return pd.concat(frames, ignore_index=True)


def _dedupe_by_docket(df: pd.DataFrame) -> pd.DataFrame:
    """
    Re-running a date that's already in the output produces records with the
    same docket_number as last time. Keep the newest version of each docket
    number (this run's), so re-scraping a date replaces its rows instead of
    duplicating them. Records with no docket number are kept as-is — they're
    never treated as duplicates of each other.
    """
    docket = df["docket_number"].astype(str).str.strip()
    has_docket = docket != ""

    with_docket = df[has_docket].copy()
    with_docket["docket_number"] = docket[has_docket]
    with_docket = with_docket.drop_duplicates(subset="docket_number", keep="last")

    without_docket = df[~has_docket]

    return pd.concat([with_docket, without_docket], ignore_index=True)


def _resolve_writable_path(filepath: str) -> str:
    """
    Check that filepath can be written to. If it's locked — most commonly
    because the user still has last run's output open in Excel — fall back
    to a timestamped filename in the same folder so the run finishes and no
    data is lost, instead of crashing.
    """
    try:
        with open(filepath, "a"):
            pass
        return filepath
    except OSError:
        base, ext = os.path.splitext(filepath)
        fallback = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        logger.warning(
            "Could not write to %s — it's likely open in Excel. "
            "Saving this run's results to %s instead. Close %s and the "
            "next run will use the normal filename again.",
            filepath, fallback, filepath,
        )
        return fallback


def _write_dataframe_to_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))


def _format_sheet(ws, df: pd.DataFrame) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
