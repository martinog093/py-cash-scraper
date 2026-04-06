"""
Output: CSV file + Excel file + Google Sheets append.
"""

import logging
import os
from datetime import date

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    "market",
    "buyer_name",
    "entity_type",
    "property_address",
    "sale_date",
    "purchase_price",
    "deed_type",
    "record_number",
    "buyer_mailing_address",
    "times_bought_90d",
    "priority",
]

PRIORITY_COLORS = {
    "Hot":      "FF4444",
    "Warm":     "FFA500",
    "Standard": "FFFFFF",
}


def write_csv(records: list[dict], output_dir: str = "output") -> str:
    """Write records to a dated CSV file. Returns the file path."""
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"cash_buyers_{date.today().isoformat()}.csv")

    df = _to_dataframe(records)
    df.to_csv(filepath, index=False)
    logger.info("CSV written: %s (%d rows)", filepath, len(df))
    return filepath


def write_excel(records: list[dict], output_dir: str = "output") -> str:
    """Write records to a formatted Excel file. Returns the file path."""
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"cash_buyers_{date.today().isoformat()}.xlsx")

    df = _to_dataframe(records)
    df.to_excel(filepath, index=False, sheet_name="Cash Buyers")

    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Bold + grey header row, freeze it
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    # Colour-code priority column (last column) and auto-fit widths
    priority_col_idx = COLUMNS.index("priority") + 1
    col_widths = {i: len(col) for i, col in enumerate(COLUMNS, start=1)}

    for row in ws.iter_rows(min_row=2):
        priority_cell = row[priority_col_idx - 1]
        priority = str(priority_cell.value or "")
        color = PRIORITY_COLORS.get(priority)
        if color:
            priority_cell.fill = PatternFill("solid", fgColor=color)
            if priority == "Hot":
                priority_cell.font = Font(bold=True, color="FFFFFF")

        for i, cell in enumerate(row, start=1):
            val_len = len(str(cell.value or ""))
            if val_len > col_widths.get(i, 0):
                col_widths[i] = val_len

    for i, width in col_widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 50)

    wb.save(filepath)
    logger.info("Excel written: %s (%d rows)", filepath, len(df))
    return filepath


def append_to_google_sheet(records: list[dict]) -> None:
    """Append records to the configured Google Sheet."""
    import gspread
    from google.oauth2.service_account import Credentials
    from dotenv import load_dotenv

    load_dotenv()
    sheet_id = os.getenv("GOOGLE_SHEET_ID", "")
    creds_file = os.getenv("GOOGLE_CREDENTIALS_FILE", "credentials/google_service_account.json")

    if not sheet_id or not os.path.exists(creds_file):
        logger.warning("Google Sheets not configured — skipping upload")
        return

    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(creds_file, scopes=scopes)
    gc = gspread.authorize(creds)
    sheet = gc.open_by_key(sheet_id).sheet1

    df = _to_dataframe(records)

    # Insert header at row 1 if it isn't already there (handles a truly
    # empty sheet and a sheet that already has data rows but no header).
    existing = sheet.get_all_values()
    if not existing or existing[0] != COLUMNS:
        sheet.insert_row(COLUMNS, index=1)

    rows_to_append = [list(row) for _, row in df.iterrows()]
    sheet.append_rows(rows_to_append)

    logger.info("Appended %d rows to Google Sheet %s", len(df), sheet_id)


def _to_dataframe(records: list[dict]) -> pd.DataFrame:
    rows = []
    for r in records:
        rows.append({col: r.get(col, "") for col in COLUMNS})
    df = pd.DataFrame(rows, columns=COLUMNS)
    # Sort: Hot first, then Warm, then Standard; within each by times_bought_90d desc
    priority_order = {"Hot": 0, "Warm": 1, "Standard": 2}
    df["_priority_sort"] = df["priority"].map(priority_order).fillna(3)
    df = df.sort_values(["_priority_sort", "times_bought_90d"], ascending=[True, False])
    df = df.drop(columns=["_priority_sort"])
    return df.reset_index(drop=True)
