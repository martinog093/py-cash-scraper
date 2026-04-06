"""
Pre-flight checks for the Memphis Scraper (Tax Lien + Probate + Divorce).

Run this BEFORE filling in real credentials to confirm:
  1. All imports resolve correctly
  2. Output module writes a valid Excel file
  3. Assessor portal is reachable and returns a parseable response
  4. Shelby County docket report (Probate) is reachable
  5. Playwright/Chromium launches correctly

Usage:
    python test_setup.py
"""

import sys
import os
import traceback

PASS = "[PASS]"
FAIL = "[FAIL]"
SKIP = "[SKIP]"


def check(label: str, fn):
    try:
        fn()
        print(f"  {PASS}  {label}")
        return True
    except Exception as e:
        print(f"  {FAIL}  {label}")
        print(f"         {e}")
        return False


def test_imports():
    print("\n[1] Import checks")

    def _imports():
        from src.auth import login
        from src.assessor import verify_ownership, search_by_name
        from src.output import write_output
        from src.scrapers.tax_lien import scrape_tax_liens
        from src.scrapers.probate import scrape_probate
        from src.scrapers.divorce import scrape_divorce
        from src.docket_lookup import get_probate_parties

    check("All modules import without error", _imports)


def test_env():
    print("\n[2] .env file")

    def _env_exists():
        assert os.path.exists(".env"), ".env file not found in current directory"

    def _env_has_keys():
        from dotenv import dotenv_values
        vals = dotenv_values(".env")
        assert "MDN_USERNAME" in vals, "MDN_USERNAME missing from .env"
        assert "MDN_PASSWORD" in vals, "MDN_PASSWORD missing from .env"

    def _env_not_empty():
        from dotenv import dotenv_values
        vals = dotenv_values(".env")
        u = vals.get("MDN_USERNAME", "")
        p = vals.get("MDN_PASSWORD", "")
        placeholders = {"YOUR_EMAIL", "YOUR_PASSWORD", "", "your_username_here", "your_password_here"}
        if u in placeholders or p in placeholders:
            print(f"         (credentials are still placeholders — fill in .env before a full run)")
        else:
            print(f"         (credentials appear to be set)")

    check(".env file exists", _env_exists)
    check(".env contains MDN_USERNAME and MDN_PASSWORD keys", _env_has_keys)
    _env_not_empty()


def test_output_module():
    print("\n[3] Output module — write dummy Excel")

    def _write():
        from src.output import write_output
        dummy = [
            {
                "filing_date": "2026-06-05",
                "record_type": "Tax Lien",
                "primary_name": "John Doe",
                "secondary_name": "",
                "docket_number": "CT-2026-001",
                "verified_address": "123 Main St Memphis TN",
                "parcel_id": "012345-67890",
                "debt_amount": "$4,500.00",
                "status": "Verified",
                "discard_reason": "",
            },
            {
                "filing_date": "2026-06-05",
                "record_type": "Tax Lien",
                "primary_name": "Jane Smith",
                "secondary_name": "",
                "docket_number": "CT-2026-002",
                "verified_address": "",
                "parcel_id": "",
                "debt_amount": "$1,200.00",
                "status": "Discarded",
                "discard_reason": "No property found under name",
            },
        ]
        out_path = "output/test_output.xlsx"
        write_output(dummy, out_path)
        assert os.path.exists(out_path), "Output file was not created"
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        assert "Verified Leads" in wb.sheetnames, "Missing 'Verified Leads' sheet"
        assert "Discarded" in wb.sheetnames, "Missing 'Discarded' sheet"
        os.remove(out_path)

    check("Writes Excel with Verified Leads + Discarded sheets", _write)


def test_assessor_portal():
    print("\n[4] Assessor portal connectivity")

    def _reachable():
        import httpx
        resp = httpx.get("https://assessormelvinburgess.com", timeout=15, follow_redirects=True)
        assert resp.status_code < 500, f"Assessor portal returned HTTP {resp.status_code}"

    def _search_endpoint():
        import httpx
        resp = httpx.get(
            "https://assessormelvinburgess.com/propertySearch",
            timeout=15,
            follow_redirects=True,
        )
        assert resp.status_code < 500, f"Search endpoint returned HTTP {resp.status_code}"
        print(f"         (content-type: {resp.headers.get('content-type', 'unknown')})")
        print(f"         (response length: {len(resp.text)} chars)")

    check("Assessor portal is reachable", _reachable)
    check("/propertySearch endpoint responds", _search_endpoint)


def test_playwright():
    print("\n[5] Playwright / Chromium")

    def _launch():
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto("about:blank")
            browser.close()

    check("Chromium launches and opens a page", _launch)


def test_docket_report():
    print("\n[6] Shelby County docket report connectivity (Probate)")

    def _reachable():
        from playwright.sync_api import sync_playwright
        from src.docket_lookup import DOCKET_SETUP_URL, USER_AGENT

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page(
                user_agent=USER_AGENT,
                extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
            )
            page.goto(
                f"{DOCKET_SETUP_URL}?case_id=PR000000&begin_date=&end_date=",
                wait_until="load",
                timeout=45000,
            )
            assert page.frame(name="Big") is not None, (
                "Docket report search form ('Big' frame) not found — "
                "site may be blocking automated access"
            )
            browser.close()

    check("Docket report search page loads (not blocked by Cloudflare)", _reachable)


def main():
    print("=" * 55)
    print("  Memphis Scraper — Setup Verification")
    print("=" * 55)

    test_imports()
    test_env()
    test_output_module()
    test_assessor_portal()
    test_playwright()
    test_docket_report()

    print("\n" + "=" * 55)
    print("  Done. Fix any [FAIL] items before running the scrapers.")
    print("  Then fill in real credentials in .env and run:")
    print("    python tax_lien_scraper.py  (Tax Liens)")
    print("    python probate_scraper.py   (Probate)")
    print("    python divorce_scraper.py   (Divorce)")
    print("=" * 55)


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
