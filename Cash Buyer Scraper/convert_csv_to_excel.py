"""
One-time utility: convert an existing cash_buyers CSV to a formatted Excel file.

Usage:
    python convert_csv_to_excel.py output\cash_buyers_2026-07-01.csv
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PRIORITY_COLORS = {
    "Hot":      "FF4444",
    "Warm":     "FFA500",
    "Standard": "FFFFFF",
}


def convert(csv_path: str) -> str:
    if not os.path.exists(csv_path):
        print(f"ERROR: file not found: {csv_path}")
        sys.exit(1)

    xlsx_path = csv_path.replace(".csv", ".xlsx")
    df = pd.read_csv(csv_path)
    df.to_excel(xlsx_path, index=False, sheet_name="Cash Buyers")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    cols = list(df.columns)
    priority_col_idx = cols.index("priority") + 1 if "priority" in cols else None
    col_widths = {i: len(str(col)) for i, col in enumerate(cols, start=1)}

    for row in ws.iter_rows(min_row=2):
        if priority_col_idx:
            priority_cell = row[priority_col_idx - 1]
            priority = str(priority_cell.value or "")
            color = PRIORITY_COLORS.get(priority)
            if color:
                priority_cell.fill = PatternFill("solid", fgColor=color)
                if priority == "Hot":
                    priority_cell.font = Font(bold=True, color="FFFFFF")

        for i, cell in enumerate(row, start=1):
            val_len = len(str(cell.value or ""))
            if val_len > col_widths.get(i, 0):
                col_widths[i] = val_len

    for i, width in col_widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 50)

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
    return xlsx_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert_csv_to_excel.py <path_to_csv>")
        sys.exit(1)
    convert(sys.argv[1])
